import streamlit as st
import requests
import json
import re
import graphviz
import io
from openpyxl import Workbook
from openpyxl.styles import Font

# API URLs
Component_Function = "https://hemapriyadharshini-flowise.hf.space/api/v1/prediction/b83d91c5-7db1-4426-815b-d3b4e48cde31"
Engineering_Requirements = "https://hemapriyadharshini-flowise.hf.space/api/v1/prediction/f9893e95-2172-4b00-9b45-b14748ad1ae9"
Component_Diagram = "https://hemapriyadharshini-flowise.hf.space/api/v1/prediction/3e98e228-996a-4205-8c75-3cf34774b088"
Function_Diagram = "https://hemapriyadharshini-flowise.hf.space/api/v1/prediction/9297a1fe-c925-4c00-9217-608d891b81ef"
FMEA = "https://hemapriyadharshini-flowise.hf.space/api/v1/prediction/fe1c2f99-b3a6-4b2f-9145-c6e07c686483"
DVP_AND_API = "https://hemapriyadharshini-flowise.hf.space/api/v1/prediction/618993c3-15d8-46b2-915d-71130090f6d1"

# Function to query Flowise API
def query_flowise_api(api_url, payload):
    response = requests.post(api_url, json=payload)
    try:
        return response.json()
    except json.JSONDecodeError:
        return {"error": "Invalid response from the API"}

# Function to convert JSON response to table data for components and functions
def json_to_table_data(response_data):
    json_str = response_data.get('text', '{}').strip('```json').strip()  # Strip Markdown and extra spaces
    
    try:
        json_data = json.loads(json_str)
    except json.JSONDecodeError:
        return [], ["Error decoding JSON from 'text' field."]
    
    components_and_functions = json_data.get('components_and_functions', [])
    
    if not components_and_functions:
        return [], ["No data found in the 'components_and_functions' key."]
    
    table_data = [[item['specification_category'], item['component'], item['function']] 
                  for item in components_and_functions]
    
    return table_data

# Function to convert JSON response to table data for Engineering Requirements
def json_to_engineering_req_table(response_data):
    text = response_data.get('text', '')
    json_match = re.search(r'```json\n([\s\S]*?)\n```', text)
    
    if not json_match:
        return [["No JSON data found in the response."]]
    
    json_str = json_match.group(1)
    
    try:
        json_data = json.loads(json_str)
    except json.JSONDecodeError:
        return [["Error decoding JSON from 'text' field."]]
    
    table_data = [[item.get('id', ''), item.get('user_need', ''), 
                   item.get('tech_id', ''), item.get('description', '')] 
                  for item in json_data]
    
    return table_data if table_data else [["No data found in the JSON."]]

# Function to convert FMEA response to table data
def json_to_fmea_table(response_data):
    json_str = response_data.get('text', '').strip('```json').strip()
    
    try:
        json_data = json.loads(json_str)
    except json.JSONDecodeError:
        return [], ["Error decoding JSON from 'text' field."]
    
    fmea_entries = json_data.get('fmea', [])
    
    if not fmea_entries:
        return [], ["No data found in the 'fmea' key."]
    
    table_data = [
        [
            entry.get('id', ''),
            entry.get('name_of_related_function', ''),
            entry.get('name_of_requirement', ''),
            entry.get('potential_failure_mode', ''),
            entry.get('effects_of_failure', ''),
            entry.get('severity', ''),
            entry.get('cause_of_failure', ''),
            entry.get('occurance', ''),
            entry.get('design_controls', ''),
            entry.get('detection', ''),
            entry.get('rpn', ''),
            entry.get('recommended_actions', '')
        ]
        for entry in fmea_entries
    ]
    
    return table_data

# Function to query the DVP&R API
def query_dvp_and_api(engineering_requirements, fmea):
    combined_question = f"Generate DVP&R based on Engineering Requirements: {engineering_requirements} and FMEA: {fmea}."
    response = requests.post(DVP_AND_API, json={"question": combined_question})
    try:
        return response.json()
    except json.JSONDecodeError:
        return {"error": "Invalid response from DVP&R API"}

# Function to convert DVP&R response to table data
def json_to_dvpr_table(response_data):
    dvpr_data = response_data.get('text', '')

    try:
        dvpr_json = json.loads(dvpr_data.strip('```json').strip())  # Parse the JSON if it's wrapped in code blocks
    except json.JSONDecodeError:
        return [["Error decoding JSON from DVP&R response."]]
    
    dvpr_table = [
        [
            item.get('related_id', ''),
            item.get('test_no', ''),
            item.get('test_name', ''),
            item.get('method', ''),
            item.get('duration',''),
            item.get('acceptance_criteria','')

        ]
        for item in dvpr_json.get('dvpr', [])
    ]
    
    return dvpr_table if dvpr_table else [["No DVP&R data found."]]

def mermaid_to_graphviz(mermaid_code):
    # Convert Mermaid flowchart to Graphviz DOT language
    lines = mermaid_code.split('\n')
    dot_code = 'digraph {\n'
    for line in lines[1:]:  # Skip the first line (flowchart TD)
        if '-->' in line:
            parts = line.split('-->')
            source = parts[0].strip().replace('[', '').replace(']', '')
            target = parts[1].strip().replace('[', '').replace(']', '')
            dot_code += f'    "{source}" -> "{target}"\n'
    dot_code += '}'
    return dot_code

# New function to create Excel file
def create_excel_file(user_input, components_functions, eng_req, fmea, dvpr):
    wb = Workbook()
    
    # Prompt sheet
    ws_prompt = wb.active
    ws_prompt.title = "Prompt"
    ws_prompt['A1'] = "Prompt"
    ws_prompt['A1'].font = Font(bold=True)
    ws_prompt['A2'] = user_input

    # Components & Functions sheet
    ws_cf = wb.create_sheet("Components & Functions")
    headers = ['Specification Category', 'Component', 'Function']
    ws_cf.append(headers)
    for row in components_functions:
        ws_cf.append(row)

    # Engineering Requirements sheet
    ws_er = wb.create_sheet("Engineering Requirements")
    headers = ['User Req ID', 'User Need', 'Technical Req ID', 'Description']
    ws_er.append(headers)
    for row in eng_req:
        ws_er.append(row)

    # FMEA sheet
    ws_fmea = wb.create_sheet("FMEA")
    headers = ['ID', 'Related Function', 'Requirement', 'Failure Mode', 'Effects', 'Severity', 
               'Cause', 'Occurrence', 'Controls', 'Detection', 'RPN', 'Recommended Actions']
    ws_fmea.append(headers)
    for row in fmea:
        ws_fmea.append(row)

    # DVP&R sheet
    ws_dvpr = wb.create_sheet("DVP&R")
    headers = ['Related ID', 'Test Number', 'Test Name', 'Test Method', 'Duration', 'Acceptance Criteria']
    ws_dvpr.append(headers)
    for row in dvpr:
        ws_dvpr.append(row)

    # Save to a bytes buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

# Streamlit app
st.title("System Design AI")

# Initialize session state to store generated table data
if 'table_data' not in st.session_state:
    st.session_state.table_data = None
    st.session_state.eng_req_data = None
    st.session_state.locked = False  # To lock the table after clicking Next
    st.session_state.component_diagram = None
    st.session_state.function_diagram = None
    st.session_state.fmea_data = None
    st.session_state.dvp_and_data = None  # To store DVP&R data

# Text input for user to enter their question
user_input = st.text_input("Generate experimentation plan for...")

# Button to generate response from the first API
if st.button("Generate") and not st.session_state.locked:  # Disable button when table is locked
    if user_input:
        with st.spinner("Generating Components and Functions..."):
            response = query_flowise_api(Component_Function, {"question": user_input})
            if "error" in response:
                st.error(response["error"])
            else:
                table_data = json_to_table_data(response)
                st.session_state.table_data = table_data

# Display the generated table if available
if st.session_state.table_data:
    st.subheader("Components and Functions:")
    headers = ['Specification Category', 'Component', 'Function']
    table_data = [headers] + st.session_state.table_data
    st.table(table_data)

    # Show the "Next" button only if the table is generated
    if st.button("Next") and not st.session_state.locked:
        st.session_state.locked = True
        question_data = {
            "components_and_functions": [
                {"specification_category": row[0], "component": row[1], "function": row[2]}
                for row in st.session_state.table_data
            ]
        }
        question_json = json.dumps(question_data)
        
        with st.spinner("Generating Engineering Requirements..."):
            response = query_flowise_api(Engineering_Requirements, {"question": question_json})
            eng_req_table_data = json_to_engineering_req_table(response)
            st.session_state.eng_req_data = eng_req_table_data

# Display the Engineering Requirements table if available
if st.session_state.eng_req_data:
    st.subheader("Engineering Requirements:")
    headers = ['User Req ID', 'User Need', 'Technical Req ID', 'Description']
    eng_req_table = [headers] + st.session_state.eng_req_data
    st.table(eng_req_table)
    
    # Add another "Next" button for the component diagram API call
    if st.button("Next for Component Diagram"):
        with st.spinner("Generating Component Diagram..."):
            response = query_flowise_api(Component_Diagram, {"question": user_input})
            if "error" in response:
                st.error(response["error"])
            else:
                st.session_state.component_diagram = response

# Display the component diagram if available
if st.session_state.component_diagram:
    st.subheader("Component Diagram:")
    component_text = st.session_state.component_diagram['text']
    # Remove the ```json prefix and ``` suffix if present
    component_text = component_text.strip().lstrip('```json').rstrip('```')
    try:
        component_json = json.loads(component_text)
        component_mermaid = component_json['mermaid_diagram']
        
        # Display raw response
        st.subheader("Raw Component Diagram Response:")
        st.code(component_mermaid, language='mermaid')
        
        # Display rendered diagram
        st.subheader("Rendered Component Diagram:")
        dot_code = mermaid_to_graphviz(component_mermaid)
        graph = graphviz.Source(dot_code)
        st.graphviz_chart(graph)
    except json.JSONDecodeError as e:
        st.error(f"Error parsing component diagram JSON: {e}")
    except Exception as e:
        st.error(f"Error displaying component diagram: {e}")
    
    # Add a "Next" button to generate the function diagram
    if st.button("Next for Function Diagram"):
        with st.spinner("Generating Function Diagram..."):
            response = query_flowise_api(Function_Diagram, {"question": user_input})
            if "error" in response:
                st.error(response["error"])
            else:
                st.session_state.function_diagram = response

# Display the function diagram if available
if st.session_state.function_diagram:
    st.subheader("Function Diagram:")
    function_text = st.session_state.function_diagram['text']
    # Remove the ```json prefix and ``` suffix if present
    function_text = function_text.strip().lstrip('```json').rstrip('```')
    try:
        function_json = json.loads(function_text)
        function_mermaid = function_json['mermaid_diagram']
        
        # Display raw response
        st.subheader("Raw Function Diagram Response:")
        st.code(function_mermaid, language='mermaid')
        
        # Display rendered diagram
        st.subheader("Rendered Function Diagram:")
        dot_code = mermaid_to_graphviz(function_mermaid)
        graph = graphviz.Source(dot_code)
        st.graphviz_chart(graph)
    except json.JSONDecodeError as e:
        st.error(f"Error parsing function diagram JSON: {e}")
    except Exception as e:
        st.error(f"Error displaying function diagram: {e}")

    # Add a "Next" button to generate the FMEA table
    if st.button("Next for FMEA"):
        with st.spinner("Generating FMEA..."):
            response = query_flowise_api(FMEA, {"question": user_input})
            if "error" in response:
                st.error(response["error"])
            else:
                fmea_table_data = json_to_fmea_table(response)
                st.session_state.fmea_data = fmea_table_data

# Display the FMEA table if available
if st.session_state.fmea_data:
    st.subheader("Failure Mode Analysis:")
    fmea_headers = ['ID', 'Related Function', 'Requirement', 'Failure Mode', 'Effects', 'Severity', 
                    'Cause', 'Occurrence', 'Controls', 'Detection', 'RPN', 'Recommended Actions']
    fmea_table = [fmea_headers] + st.session_state.fmea_data
    st.table(fmea_table)
    
    # Add a "Next" button for the DVP&R table generation
    if st.button("Generate DVP&R"):
        with st.spinner("Generating DVP&R..."):
            dvpr_response = query_dvp_and_api(st.session_state.eng_req_data, st.session_state.fmea_data)
            dvpr_table_data = json_to_dvpr_table(dvpr_response)
            st.session_state.dvp_and_data = dvpr_table_data

# Display the DVP&R table if available
if st.session_state.dvp_and_data:
    st.subheader("Design Verification Plan:")
    dvpr_headers = ['Related ID', 'Test Number', 'Test Name', 'Test Method', 'Duration', 'Acceptance Criteria']
    dvpr_table = [dvpr_headers] + st.session_state.dvp_and_data
    st.table(dvpr_table)

    # Add Download Plan button
    if st.button("Download Plan"):
        excel_buffer = create_excel_file(
            user_input,
            st.session_state.table_data,
            st.session_state.eng_req_data,
            st.session_state.fmea_data,
            st.session_state.dvp_and_data
        )
        st.download_button(
            label="Download Excel File",
            data=excel_buffer,
            file_name="system_design_plan.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )